import pandas as pd

from flask import (
    Blueprint,
    render_template,
    request,
    jsonify
)

from search.module.data_loader import load_data
from search.module.kpi import calculate_kpi, calculate_report
from search.module.search import apply_search
from search.module.search_help import HELP_SECTIONS
from common.utils.parsers import (
    format_display_date
)
from common.utils.links import (
    get_url
)

from common.utils.user_group import (
    save_user_group,
    filter_dataframe_by_group
)

from search.module.data_loader import (
    load_group_filters,
    load_group_users
)

# ── Logger ────────────────────────────────────────────────────────────────────
try:
    from common.logger import setup_logger
    logger = setup_logger("search")
except Exception:
    import logging
    logger = logging.getLogger("search")



# -----------------------------------
# Blueprint
# -----------------------------------
search_bp = Blueprint(
    "search",
    __name__,
    template_folder="templates",
    static_folder="statics",
    static_url_path="/search/static"
)


# -----------------------------------
# Search Page
# -----------------------------------
@search_bp.route("/search")
def search_page():
    try:
        logger.info("Search page loaded")
        df, last_refresh = load_data()
        kpi = calculate_kpi(df)
        logger.info(f"Data loaded — {len(df)} records | KPI total={kpi.get('total',0)}")

        return render_template(
            "search.html",
            last_refresh=last_refresh,
            kpi=kpi
        )

    except Exception as e:
        logger.error(f"Search page error: {e}", exc_info=True)
        return str(e)


# -----------------------------------
# Filter Options API
# -----------------------------------
@search_bp.route("/search/filter-options")
def search_filter_options():

    try:
        logger.debug("Filter options requested")
        df, _ = load_data()

        status = sorted(
            df["Status"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        priority = sorted(
            df["Priority"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        groups = load_group_filters()

        return jsonify({
            "status": status,
            "priority": priority,
            "groups": groups
        })

    except Exception as e:

        return jsonify({
            "error": str(e)
        })

# -----------------------------------
# SAVE USER GROUP
# -----------------------------------
@search_bp.route(
    "/search/save-group",
    methods=["POST"]
)
def save_group():

    try:

        data = request.json

        group_name = data.get(
            "group_name",
            ""
        ).strip()

        users = data.get(
            "users",
            []
        )

        if not group_name:

            return jsonify({
                "success": False,
                "message": "Group name required"
            })

        logger.info(f"Saving group '{group_name}' with {len(users)} members")
        save_user_group(
            group_name,
            users
        )

        return jsonify({
            "success": True
        })

    except Exception as e:
        logger.error(f"Save group error: {e}", exc_info=True)
        return jsonify({
            "success": False,
            "message": str(e)
        })


# -----------------------------------
# Search Issues API
# -----------------------------------
@search_bp.route(
    "/search/issues",
    methods=["POST"]
)
def search_issues():
    try:

        data = request.json

        query = data.get(
            "query",
            ""
        )

        sources = data.get(
            "sources",
            []
        )

        status = data.get(
            "status",
            ""
        )

        priority = data.get(
            "priority",
            ""
        )

        group = data.get(
            "group",
            ""
        )

        search_in = data.get(
            "search_in",
            ["short_description"]
        )

        date_field = data.get(
            "date_field",
            "created"
        )

        start_date = data.get(
            "start_date",
            ""
        )

        end_date = data.get(
            "end_date",
            ""
        )

        df, _ = load_data()

        # -----------------------------------
        # SOURCE FILTER
        # -----------------------------------
        if sources:
            df = df[
                df["Source"].isin(sources)
            ]

        # -----------------------------------
        # STATUS FILTER
        # -----------------------------------
        if status:
            df = df[
                df["Status"].astype(str) == status
            ]


        # -----------------------------------
        # PRIORITY FILTER
        # -----------------------------------
        if priority:
            df = df[
                df["Priority"].astype(str) == priority
            ]


        # -----------------------------------
        # GROUP FILTER
        # -----------------------------------
        if group:

            from common.utils.user_group import (
                build_group_mapping
            )

            mapping_df = build_group_mapping(df)

            df = filter_dataframe_by_group(
                df,
                mapping_df,
                [group]
            )

        # -----------------------------------
        # DATE FIELD
        # -----------------------------------
        date_column = (
            "Created Date"
            if date_field == "created"
            else "Resolved Date"
        )


        # -----------------------------------
        # DATE RANGE FILTER
        # -----------------------------------

        if start_date:

            start_date = pd.to_datetime(start_date)

            df = df[
                df[date_column] >= start_date
            ]


        if end_date:

            end_date = pd.to_datetime(end_date)

            df = df[
                df[date_column] <= end_date
            ]

        # -----------------------------------
        # SEARCH
        # -----------------------------------
        filtered = apply_search(
            df,
            query,
            search_in=search_in
        )

        filtered = filtered.fillna("")

        results = []

        for _, row in filtered.iterrows():

            source = row.get(
                "Source",
                ""
            )

            number = str(
                row.get(
                    "Number",
                    ""
                )
            )

            # -----------------------------
            # external links
            # -----------------------------
            if source == "SNOW":
                url = get_url("incident", number)

            elif source == "PTC":
                url = get_url("ptc case", number)

            elif source == "AZURE":
                url = get_url("azure bug", number)

            elif source == "AOM":
                url = get_url("azure_vpa", number)

            else:
                url = ""

            results.append({
                "number": number,
                "description": row.get(
                    "Description",
                    ""
                ),
                "vendor_ticket":    str(row.get("Vendor Ticket",    "") or ""),
                "azure_bug":        str(row.get("Azure Bug",        "") or ""),
                "azure_user_story": str(row.get("Azure User Story", "") or ""),
                "ptc_articles":     str(row.get("PTC Articles",     "") or ""),
                "priority": row.get(
                    "Priority",
                    ""
                ),
                "status": row.get(
                    "Status",
                    ""
                ),
                "created_by": row.get(
                    "Created By",
                    ""
                ),
                "created_date": format_display_date(
                    row.get("Created Date")
                ),
                "assigned_to": row.get(
                    "Assigned To",
                    ""
                ),
                "resolved_date": format_display_date(
                    row.get("Resolved Date")
                ),
                "source": source,
                "url": url
            })

        _counts = {}
        for _r in results:
            _counts[_r.get("source","?")] = _counts.get(_r.get("source","?"), 0) + 1
        logger.info(
            f"Search — query={repr(query)} sources={sources} "
            f"status={repr(status)} priority={repr(priority)} "
            f"group={repr(group)} search_in={search_in} "
            f"results={len(results)} breakdown={_counts}"
        )

        return jsonify({
            "results": results
        })

    except Exception as e:
        logger.error(f"Search issues error: {e}", exc_info=True)
        return jsonify({
            "error": str(e)
        })


@search_bp.route("/search/group-members")
def get_group_members():

    try:

        from common.utils.user_group import (
            load_group_mapping
        )

        mapping = load_group_mapping()

        grouped = {}

        for group_name, group_df in mapping.groupby("Group"):

            grouped[group_name] = sorted(
                group_df["Name"]
                .dropna()
                .astype(str)
                .tolist()
            )

        return jsonify({
            "groups": grouped
        })

    except Exception as e:

        return jsonify({
            "error": str(e)
        })

@search_bp.route("/search/group-users")
def get_group_users():

    try:

        users = load_group_users()

        return jsonify({
            "users": users
        })

    except Exception as e:

        return jsonify({
            "users": [],
            "error": str(e)
        }), 500

# -----------------------------------
# COLLECT USERS FROM RAW DATA FILES
# -----------------------------------
@search_bp.route("/search/collect-users", methods=["POST"])
def collect_users_from_data():
    """
    Reads Azure.csv (Created By), Snow.xlsx (Assigned To), Ptc.csv (Case Contact),
    parses out clean display names (strips emails/IDs), merges with existing
    user_group_mapping.csv preserving group assignments, saves back, returns user list.
    """
    import re
    import pandas as pd
    from pathlib import Path

    EMAIL_RE        = re.compile(r'[\w\.\+\-]+@[\w\.\-]+\.\w+')
    ANGLE_RE        = re.compile(r'<[^>]+>')
    PARENS_EMAIL_RE = re.compile(r'\([^)]*@[^)]*\)')
    PURE_ID_RE      = re.compile(r'^[A-Za-z]{1,6}\d+$')
    ALL_DIGIT_RE    = re.compile(r'^\d+$')
    DOMAIN_RE       = re.compile(r'[A-Za-z0-9_\-]+\\[A-Za-z0-9_\-]+')
    SYSTEM_ACCOUNT_RE = re.compile(r'^[A-Z][a-z]+(?:[A-Z][a-z0-9]*){2,}$')

    # Role/label words inside parentheses — strip them
    ROLE_WORDS = {
        "consultant", "contractor", "admin", "administrator",
        "system", "service", "integration", "automation",
        "robot", "bot", "internal", "external",
        "guest", "temp", "temporary"
    }

    # Words that flag a value as a group/queue/team, not a person
    GROUP_KEYWORDS = {
        "support", "team", "group", "queue", "level", "tier",
        "helpdesk", "service", "desk", "operations", "ops",
        "department", "dept", "division", "center", "centre",
        "unassigned", "unknown", "nobody", "none", "n/a",
        "shared", "generic", "global", "regional", "local",
        "escalation", "on-call", "oncall", "shift",
        "management", "leadership", "engineering", "infrastructure",
        "network", "security", "monitoring", "maintenance",
    }

    def is_id_token(token):
        t = token.strip()
        if not t:
            return True
        if ALL_DIGIT_RE.match(t):
            return True
        if PURE_ID_RE.match(t):
            return True
        return False

    def is_group_name(s):
        """
        Return True if s looks like a team/queue/group name rather than a person.
        Heuristics:
          - Contains a digit-ordinal prefix: "2nd", "3rd", "1st" …
          - Contains a group keyword as a standalone word
          - Starts with a digit (e.g. "2nd Level Support")
        """
        lower_tokens = [t.lower().strip('.,;') for t in s.split()]
        # Starts with a number-word like "2nd", "3rd", "1st", "2"
        if lower_tokens and re.match(r'^\d', lower_tokens[0]):
            return True
        # Contains any group keyword
        if any(t in GROUP_KEYWORDS for t in lower_tokens):
            return True
        return False

    def strip_role_parens(m):
        inner = m.group(1).strip().lower()
        return "" if inner in ROLE_WORDS else m.group(0)

    def clean_name(raw):
        """
        Return a clean human display name, or None if the value is an
        ID, email, system account, group name, or otherwise non-personal.
        """
        if not raw or str(raw).strip() in ("", "nan", "NaN"):
            return None

        s = str(raw).strip()

        # Strip leading/trailing punctuation that sometimes prefixes names: ". Karthikeyan"
        s = re.sub(r'^[\.\,\;\:\-\s]+', '', s).strip()
        s = re.sub(r'[\.\,\;\:\-\s]+$', '', s).strip()

        if not s:
            return None

        # Remove DOMAIN\user patterns
        s = DOMAIN_RE.sub("", s)

        # Remove angle-bracket content
        s = ANGLE_RE.sub("", s)

        # Remove parens with emails
        s = PARENS_EMAIL_RE.sub("", s)

        # Remove parens containing a digit (ID references): "(a447927: Volvo...)"
        s = re.sub(r'\([^)]*\d[^)]*\)', '', s)

        # Remove parens with pure role words: "(Consultant)"
        s = re.sub(r'\(([^)]+)\)', strip_role_parens, s)

        # Skip pure email values
        if EMAIL_RE.fullmatch(s.strip()):
            return None
        s = EMAIL_RE.sub("", s)

        # Clean stray punctuation
        s = re.sub(r'[\(\)\[\]<>]', '', s)
        s = re.sub(r'[:,;]+', ' ', s)
        s = re.sub(r'\s+', ' ', s).strip().strip('-').strip(',').strip()
        # Strip leading dots/punctuation again after all replacements
        s = re.sub(r'^[\.\,\;\:\-\s]+', '', s).strip()

        if not s:
            return None

        # Single-word checks — no space means likely a code, ID or system account
        if ' ' not in s:
            if is_id_token(s):
                return None
            if SYSTEM_ACCOUNT_RE.match(s) or re.search(r'\d', s):
                return None

        tokens = s.split()
        if all(is_id_token(t) for t in tokens):
            return None

        clean_tokens = [t for t in tokens if not is_id_token(t)]
        if not clean_tokens:
            return None

        result = " ".join(clean_tokens).strip()

        if len(result) < 2 or not re.search(r'[A-Za-z]', result):
            return None

        # Must have at least one word with 2+ letters
        if not any(re.match(r'[A-Za-zÀ-ÿ]{2,}', t) for t in clean_tokens):
            return None

        # Reject group/team/queue names
        if is_group_name(result):
            return None

        return result

    users = set()

    try:
        azure = pd.read_csv("data/Azure.csv")
        azure.columns = azure.columns.str.strip().str.lower()
        for col_name in ["created by", "assigned to"]:
            if col_name in azure.columns:
                for v in azure[col_name].dropna():
                    n = clean_name(v)
                    if n:
                        users.add(n)
    except Exception:
        pass

    try:
        snow = pd.read_excel("data/Snow.xlsx", engine="openpyxl")
        snow.columns = snow.columns.str.strip().str.lower()
        for col_name in ["assigned to", "opened by", "created by"]:
            if col_name in snow.columns:
                for v in snow[col_name].dropna():
                    n = clean_name(v)
                    if n:
                        users.add(n)
    except Exception:
        pass

    try:
        ptc = pd.read_csv("data/Ptc.csv", index_col=False, engine="python")
        ptc.columns = ptc.columns.str.strip().str.lower()
        for col_name in ["case contact", "case assignee", "created by"]:
            if col_name in ptc.columns:
                for v in ptc[col_name].dropna():
                    n = clean_name(v)
                    if n:
                        users.add(n)
    except Exception:
        pass

    sorted_users = sorted(users)

    # Merge with existing CSV — preserve existing group assignments
    csv_path = Path("data/user_group_mapping.csv")
    if csv_path.exists():
        existing = pd.read_csv(csv_path)
        if "Name" not in existing.columns:
            existing["Name"] = ""
        if "Group" not in existing.columns:
            existing["Group"] = "UNASSIGNED"
        existing["Name"] = existing["Name"].astype(str).str.strip()
        existing["Group"] = existing["Group"].astype(str).str.strip()
        existing_map = dict(zip(existing["Name"], existing["Group"]))
    else:
        existing_map = {}

    rows = []
    for u in sorted_users:
        rows.append({
            "Name": u,
            "Group": existing_map.get(u, "UNASSIGNED")
        })

    pd.DataFrame(rows).to_csv(csv_path, index=False)

    logger.info(f"Collect users — {len(sorted_users)} unique names saved to {csv_path}")

    return jsonify({
        "success": True,
        "users": sorted_users,
        "count": len(sorted_users)
    })


# -----------------------------------
# HELP GUIDE DATA
# -----------------------------------

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL DOWNLOAD  —  server-side openpyxl with full formatting
# ─────────────────────────────────────────────────────────────────────────────
@search_bp.route("/search/download", methods=["POST"])
def download_excel():
    """
    Accepts the current result set as JSON, builds a fully-formatted
    openpyxl workbook (data sheet + dashboard sheet) and streams it
    back as an .xlsx attachment.
    """
    import io, re
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import ColorScaleRule, Rule
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel

    try:
        payload = request.get_json(force=True)
        rows    = payload.get("rows", [])
        logger.info(f"Download requested — {len(rows)} rows")

        # ── helpers ──────────────────────────────────────────────────────────
        def clean(v):
            if not v: return ""
            return re.sub(r"<[^>]+>", "", str(v)).strip()

        def fmt_vpa(raw):
            if not raw: return ""
            parts = []
            for entry in str(raw).split(","):
                entry = entry.strip()
                if "|" in entry:
                    sid, env = entry.split("|", 1)
                    parts.append(f"{sid} ({env})" if env else sid)
                else:
                    parts.append(entry)
            return ", ".join(parts)

        # ── STATUS palette ───────────────────────────────────────────────────
        STATUS_FILL = {
            "open":        "FFFEF3C7",   # amber
            "in progress": "FFDBEAFE",   # blue
            "on hold":     "FFFCE7F3",   # pink
            "resolved":    "FFD1FAE5",   # green
            "closed":      "FFD1FAE5",   # green
            "cancelled":   "FFF3F4F6",   # grey
            "new":         "FFEDE9FE",   # purple
        }
        STATUS_FONT = {
            "open":        "FF92400E",
            "in progress": "FF1E40AF",
            "on hold":     "FF9D174D",
            "resolved":    "FF065F46",
            "closed":      "FF065F46",
            "cancelled":   "FF374151",
            "new":         "FF4C1D95",
        }

        def status_key(s):
            return str(s).lower().strip()

        # ── STYLES ───────────────────────────────────────────────────────────
        thin  = Side(style="thin",   color="FFD1D5DB")
        thick = Side(style="medium", color="FF9CA3AF")
        border_data   = Border(left=thin, right=thin, top=thin, bottom=thin)
        border_header = Border(left=thick, right=thick, top=thick, bottom=thick)

        def hdr_style(bg="FF1E3A5F", fg="FFFFFFFF"):
            return {
                "font":      Font(bold=True, color=fg, name="Calibri", size=10),
                "fill":      PatternFill("solid", fgColor=bg),
                "alignment": Alignment(horizontal="center", vertical="center",
                                       wrap_text=True),
                "border":    border_header,
            }

        def apply(cell, **kwargs):
            for k, v in kwargs.items():
                setattr(cell, k, v)

        # ── WORKBOOK ─────────────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Search Results"

        # ── HEADERS (2-row: group label + sub-column) ────────────────────────
        MAX_MULTI = 3

        def split_ids(raw, sep=","):
            if not raw: return []
            return [x.strip() for x in str(raw).split(sep) if x.strip()]

        max_vt  = min(max((len(split_ids(r.get("vendor_ticket","")))    for r in rows), default=1), MAX_MULTI)
        max_ab  = min(max((len(split_ids(r.get("azure_bug","")))        for r in rows), default=1), MAX_MULTI)
        max_aus = min(max((len(split_ids(r.get("azure_user_story",""))) for r in rows), default=1), MAX_MULTI)
        max_ptc = min(max((len(split_ids(r.get("ptc_articles","")))     for r in rows), default=1), MAX_MULTI)

        # Column index helpers (1-based)
        VT_START   = 3
        AB_START   = VT_START  + max_vt
        AUS_START  = AB_START  + max_ab
        PTC_START  = AUS_START + max_aus
        PRI_COL    = PTC_START + max_ptc
        STATUS_COL = PRI_COL + 1
        total_cols = STATUS_COL + 4   # Priority Status CreatedBy CreatedDate AssignedTo ResolvedDate Source

        hs_hdr   = hdr_style()                          # dark navy
        hs_grp   = hdr_style("FF2D5986", "FFFFFFFF")   # slightly lighter for group row

        # ── Row 1: group labels (merged spans) ───────────────────────────────
        def write_group(col, span, label):
            if span > 1:
                ws.merge_cells(start_row=1, start_column=col,
                               end_row=1,   end_column=col + span - 1)
            cell = ws.cell(row=1, column=col, value=label)
            apply(cell, **hs_grp)

        write_group(1,        1,       "Number")
        write_group(2,        1,       "Description")
        write_group(VT_START, max_vt,  "Vendor Ticket")
        write_group(AB_START, max_ab,  "Azure Bug")
        write_group(AUS_START,max_aus, "Azure User Story")
        write_group(PTC_START,max_ptc, "PTC Article")
        write_group(PRI_COL,  1,       "Priority")
        write_group(STATUS_COL, 1,     "Status")
        write_group(STATUS_COL+1, 1,   "Created By")
        write_group(STATUS_COL+2, 1,   "Created Date")
        write_group(STATUS_COL+3, 1,   "Assigned To")
        write_group(STATUS_COL+4, 1,   "Resolved Date")
        write_group(STATUS_COL+5, 1,   "Source")

        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(PRI_COL + 6)}1"

        # STATUS_COL is now computed dynamically above

        # ── URL BUILDERS ─────────────────────────────────────────────────────
        def number_url(src, num):
            if not num: return ""
            if src == "SNOW":  return f"https://volvoitsm.service-now.com/nav_to.do?uri=incident.do?sysparm_query=number={num}"
            if src == "AZURE": return f"https://dev.azure.com/VolvoGroup-DVP/VCEWindchillPLM/_workitems/edit/{num}"
            if src == "PTC":   return f"https://support.ptc.com/appserver/cs/view/case.jsp?n={num}"
            if src == "AOM":   return f"https://dev.azure.com/VolvoGroup-DVP/VPA/_workitems/edit/{num}"
            return ""

        def vendor_url(v):
            if not v: return ""
            return f"https://support.ptc.com/appserver/cs/view/case.jsp?n={re.sub(r'^[Cc]', '', v.strip())}"

        def azure_bug_url(bid):
            return f"https://dev.azure.com/VolvoGroup-DVP/VCEWindchillPLM/_workitems/edit/{bid.strip()}"

        def vpa_url(sid):
            return f"https://dev.azure.com/VolvoGroup-DVP/VPA/_workitems/edit/{sid.strip()}"

        def ptc_article_url(aid):
            return f"https://www.ptc.com/en/support/article/{aid.strip()}"

        LINK_FONT = Font(name="Calibri", size=10, color="FF0563C1",
                         underline="single")

        def set_link(cell, url, display=None):
            """Apply hyperlink + blue underline font to a cell."""
            if not url: return
            if display is not None:
                cell.value = display
            cell.hyperlink = url
            cell.font = LINK_FONT

        # ── DATA ROWS ────────────────────────────────────────────────────────
        for ri, row in enumerate(rows, 2):  # data starts row 2 (1 header row)
            status_raw = clean(row.get("status", ""))
            sk         = status_key(status_raw)
            s_fill_hex = STATUS_FILL.get(sk, "FFFFFFFF")
            s_font_hex = STATUS_FONT.get(sk, "FF111827")
            src        = clean(row.get("source", ""))
            num        = clean(row.get("number", ""))

            # Build split multi-value lists
            vt_ids  = split_ids(clean(row.get("vendor_ticket", "")))[:max_vt]
            ab_ids  = split_ids(clean(row.get("azure_bug", "")))[:max_ab]
            aus_ids = split_ids(row.get("azure_user_story", ""))[:max_aus]
            ptc_ids = split_ids(clean(row.get("ptc_articles", "")))[:max_ptc]

            record = [num, clean(row.get("description", ""))]
            record += (vt_ids  + [""] * (max_vt  - len(vt_ids)))
            record += (ab_ids  + [""] * (max_ab  - len(ab_ids)))
            # Azure user story: "ID|ENV" → "ID (ENV)"
            aus_display = []
            for entry in (aus_ids + [""] * (max_aus - len(aus_ids))):
                if entry and "|" in entry:
                    sid, _, env = entry.partition("|")
                    aus_display.append(f"{sid} ({env})" if env else sid)
                else:
                    aus_display.append(entry)
            record += aus_display
            record += (ptc_ids + [""] * (max_ptc - len(ptc_ids)))
            record += [
                clean(row.get("priority", "")),
                status_raw,
                clean(row.get("created_by", "")),
                clean(row.get("created_date", "")),
                clean(row.get("assigned_to", "")),
                clean(row.get("resolved_date", "")),
                src,
            ]

            alt_fill = "FFF8FAFC" if ri % 2 == 0 else "FFFFFFFF"

            for ci, val in enumerate(record, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = Font(name="Calibri", size=10,
                                      color=s_font_hex if ci == STATUS_COL else "FF111827")
                cell.border    = border_data
                # Description wraps; all others center
                is_desc = (ci == 2)
                cell.alignment = Alignment(vertical="center",
                                           wrap_text=is_desc,
                                           horizontal="left" if is_desc else "center")
                if ci == STATUS_COL and sk in STATUS_FILL:
                    cell.fill = PatternFill("solid", fgColor=s_fill_hex)
                else:
                    cell.fill = PatternFill("solid", fgColor=alt_fill)

            ws.row_dimensions[ri].height = 16

            # ── HYPERLINKS — each ID in its own cell ─────────────────────────
            # Number (col 1)
            nurl = number_url(src, num)
            if nurl:
                set_link(ws.cell(row=ri, column=1), nurl)

            # Vendor Tickets
            for i, vt in enumerate(vt_ids):
                if vt:
                    set_link(ws.cell(row=ri, column=VT_START + i),
                             vendor_url(vt), display=vt)

            # Azure Bugs
            for i, bid in enumerate(ab_ids):
                if bid:
                    set_link(ws.cell(row=ri, column=AB_START + i),
                             azure_bug_url(bid), display=bid)

            # Azure User Stories — link to VPA, display "ID (ENV)"
            for i, entry in enumerate(aus_ids):
                if entry:
                    sid = entry.split("|")[0].strip()
                    display_aus = aus_display[i] if i < len(aus_display) else sid
                    if sid:
                        set_link(ws.cell(row=ri, column=AUS_START + i),
                                 vpa_url(sid), display=display_aus)

            # PTC Articles
            for i, aid in enumerate(ptc_ids):
                if aid:
                    set_link(ws.cell(row=ri, column=PTC_START + i),
                             ptc_article_url(aid), display=aid)

        # ── AUTO-FIT COLUMN WIDTHS ───────────────────────────────────────────
        # Calculate max content width per column across all rows
        col_max = {}
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value is None: continue
                val_len = len(str(cell.value))
                col_max[cell.column] = max(col_max.get(cell.column, 0), val_len)

        # Min widths per column type
        MIN_W = {1: 14, 2: 30}   # Number min 14, Description min 30
        for ci, max_len in col_max.items():
            min_w = MIN_W.get(ci, 8)
            # Clamp: min_w ≤ width ≤ 40; add 2 for padding
            width = max(min_w, min(max_len + 2, 40))
            ws.column_dimensions[get_column_letter(ci)].width = width

        # ── DASHBOARD SHEET ──────────────────────────────────────────────────
        from collections import defaultdict
        from datetime import datetime as _dt

        wd = wb.create_sheet("Dashboard")
        wd.sheet_view.showGridLines = False
        ws.sheet_view.showGridLines = True

        # ── KPI counts ───────────────────────────────────────────────────────
        total      = len(rows)
        open_cnt   = sum(1 for r in rows if re.search(r"open|new|progress|hold", str(r.get("status","")).lower()))
        closed_cnt = sum(1 for r in rows if re.search(r"closed|resolved", str(r.get("status","")).lower()))
        cancel_cnt = sum(1 for r in rows if "cancel" in str(r.get("status","")).lower())
        snow_cnt   = sum(1 for r in rows if r.get("source") == "SNOW")
        azure_cnt  = sum(1 for r in rows if r.get("source") == "AZURE")
        ptc_cnt    = sum(1 for r in rows if r.get("source") == "PTC")
        aom_cnt    = sum(1 for r in rows if r.get("source") == "AOM")

        # ── LAYOUT matching image 2 ──────────────────────────────────────────
        # Row 1:  Title (B1:I1)
        # Row 2:  KPI labels  B-C=Total  D-E=Open  F-G=Closed  H-I=Cancelled
        # Row 3:  KPI values
        # Row 4:  spacer
        # Row 5:  "Source Breakdown" label (B5)  |  "Monthly Pivot" title (E5)
        # Row 6:  Source hdr (B6:C6)             |  Pivot hdr Month/Year (E6:G6)
        # Rows 7-10: Source data (B7:C10)        |  Pivot data (E7:P18)
        # Chart anchored at K2 (right of KPIs, same row)

        # Column widths
        for col, w in [("A",1),("B",14),("C",14),("D",14),("E",14),
                        ("F",14),("G",14),("H",14),("I",14),("J",1)]:
            wd.column_dimensions[col].width = w

        # ── Title row 1 ───────────────────────────────────────────────────────
        wd.merge_cells("B1:I1")
        t = wd.cell(row=1, column=2, value="Search Results Dashboard")
        t.font      = Font(bold=True, color="FF1E3A5F", size=14, name="Calibri")
        t.alignment = Alignment(horizontal="left", vertical="center")
        wd.row_dimensions[1].height = 26

        # ── KPI cards (rows 2-3, 4×2-col cards: B-C, D-E, F-G, H-I) ────────
        KPI_CARDS = [
            ("Total",     total,      "FF1E3A5F", "FFFFFFFF"),
            ("Open",      open_cnt,   "FFD97706", "FFFFFFFF"),
            ("Closed",    closed_cnt, "FF065F46", "FFFFFFFF"),
            ("Cancelled", cancel_cnt, "FF6B7280", "FFFFFFFF"),
        ]
        for ki, (label, val, bg, fg) in enumerate(KPI_CARDS):
            cs = 2 + ki * 2   # B=2, D=4, F=6, H=8
            ce = cs + 1

            wd.merge_cells(start_row=2, start_column=cs, end_row=2, end_column=ce)
            lc = wd.cell(row=2, column=cs, value=label)
            lc.font = Font(bold=True, color=fg, name="Calibri", size=11)
            lc.fill = PatternFill("solid", fgColor=bg)
            lc.alignment = Alignment(horizontal="center", vertical="center")

            wd.merge_cells(start_row=3, start_column=cs, end_row=3, end_column=ce)
            vc = wd.cell(row=3, column=cs, value=val)
            vc.font = Font(bold=True, color=bg, name="Calibri", size=22)
            vc.fill = PatternFill("solid", fgColor="FFF9FAFB")
            vc.alignment = Alignment(horizontal="center", vertical="center")

            for r in (2, 3):
                for c in range(cs, ce + 1):
                    wd.cell(row=r, column=c).border = Border(
                        left  =Side(style="medium", color=bg),
                        right =Side(style="medium", color=bg),
                        top   =Side(style="medium", color=bg),
                        bottom=Side(style="medium", color=bg))

        wd.row_dimensions[2].height = 22
        wd.row_dimensions[3].height = 36
        wd.row_dimensions[4].height = 6   # spacer

        # ── Source breakdown (B5:C10) ─────────────────────────────────────────
        sb = wd.cell(row=5, column=2, value="Source Breakdown")
        sb.font = Font(bold=True, color="FF1E3A5F", size=11, name="Calibri")
        wd.row_dimensions[5].height = 18

        apply(wd.cell(row=6, column=2, value="Source"), **hdr_style())
        apply(wd.cell(row=6, column=3, value="Count"),  **hdr_style())
        wd.row_dimensions[6].height = 16
        wd.column_dimensions["B"].width = 12
        wd.column_dimensions["C"].width = 10

        src_rows_d = [("SNOW",  snow_cnt,"FFDBEAFE"),("AZURE",azure_cnt,"FFEDE9FE"),
                      ("PTC",   ptc_cnt, "FFFEF9C3"),("AOM",  aom_cnt,  "FFD1FAE5")]
        for si, (src, cnt, clr) in enumerate(src_rows_d, 7):
            c1 = wd.cell(row=si, column=2, value=src)
            c1.font = Font(bold=True, name="Calibri", size=10)
            c1.fill = PatternFill("solid", fgColor=clr)
            c1.alignment = Alignment(horizontal="center")
            c1.border = border_data
            c2 = wd.cell(row=si, column=3, value=cnt)
            c2.font = Font(name="Calibri", size=10)
            c2.alignment = Alignment(horizontal="center")
            c2.border = border_data
            wd.row_dimensions[si].height = 15

        # ── Monthly pivot (E5 onward, matching image 2) ──────────────────────
        MONTHS = ["Jan","Feb","Mar","Apr","May","Jun",
                  "Jul","Aug","Sep","Oct","Nov","Dec"]
        pivot = defaultdict(lambda: defaultdict(int))
        years_seen = set()

        for row in rows:
            cd = str(row.get("created_date", ""))
            for fmt in ("%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d", "%b %d, %Y", "%d-%B-%Y"):
                try:
                    dt = _dt.strptime(cd.strip(), fmt)
                    pivot[dt.month][dt.year] += 1
                    years_seen.add(dt.year)
                    break
                except: pass

        years_sorted = sorted(years_seen) if years_seen else []
        n_yr   = max(len(years_sorted), 1)
        pc     = 5    # pivot col E (=5)
        pr_piv = 5    # pivot title row

        wd.column_dimensions["D"].width = 1   # gap before pivot

        # Pivot title
        wd.merge_cells(start_row=pr_piv, start_column=pc,
                       end_row=pr_piv,   end_column=pc + n_yr)
        pt = wd.cell(row=pr_piv, column=pc, value="Monthly Pivot — Created Date")
        pt.font      = Font(bold=True, color="FF1E3A5F", size=11, name="Calibri")
        pt.alignment = Alignment(horizontal="center")

        # Pivot header (row 6)
        apply(wd.cell(row=pr_piv+1, column=pc, value="Month"), **hdr_style())
        wd.column_dimensions[get_column_letter(pc)].width = 9
        for yi, yr in enumerate(years_sorted):
            apply(wd.cell(row=pr_piv+1, column=pc+1+yi, value=yr), **hdr_style())
            wd.column_dimensions[get_column_letter(pc+1+yi)].width = 9
        wd.row_dimensions[pr_piv+1].height = 18

        # Pivot data (rows 7-18)
        for mi, mon in enumerate(MONTHS, 1):
            row_i = pr_piv + 2 + (mi - 1)
            mc = wd.cell(row=row_i, column=pc, value=mon)
            mc.font      = Font(bold=True, name="Calibri", size=10)
            mc.fill      = PatternFill("solid", fgColor="FFF1F5F9")
            mc.border    = border_data
            mc.alignment = Alignment(horizontal="center")
            for yi, yr in enumerate(years_sorted):
                cnt = pivot[mi].get(yr, 0)
                vc  = wd.cell(row=row_i, column=pc+1+yi, value=cnt)
                vc.font      = Font(name="Calibri", size=10)
                vc.border    = border_data
                vc.alignment = Alignment(horizontal="center")
                vc.fill      = PatternFill("solid", fgColor="FFEFF6FF" if cnt > 0 else "FFFFFFFF")
            wd.row_dimensions[row_i].height = 15

        # ── Bar chart — anchored at K2 (right of KPIs) ───────────────────────
        from openpyxl.chart.label import DataLabel
        from openpyxl.chart.data_source import NumDataSource, NumRef

        chart = BarChart()
        chart.type      = "col"
        chart.grouping  = "clustered"
        chart.style     = 10
        chart.width     = 22
        chart.height    = 14

        # Remove title, axes, legend — match image 6 clean look
        chart.title       = None
        chart.legend      = None
        chart.y_axis.delete = True   # hide Y axis
        chart.x_axis.delete = False

        if years_sorted:
            cats = Reference(wd, min_col=pc,
                             min_row=pr_piv+2, max_row=pr_piv+13)
            for yi in range(len(years_sorted)):
                chart.add_data(
                    Reference(wd, min_col=pc+1+yi,
                              min_row=pr_piv+1, max_row=pr_piv+13),
                    titles_from_data=True)
            chart.set_categories(cats)

        # Data labels on bars
        from openpyxl.chart.label import DataLabelList
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal       = True
        chart.dLbls.showLegendKey = False
        chart.dLbls.showCatName   = False
        chart.dLbls.showSerName   = False
        chart.dLbls.showPercent   = False

        # Data table below chart
        chart.plotVisOnly = True
        try:
            from openpyxl.chart.plotarea import DataTable
            chart.plot_area.dTable = DataTable()
            chart.plot_area.dTable.showHorzBorder = True
            chart.plot_area.dTable.showVertBorder = True
            chart.plot_area.dTable.showOutline    = True
            chart.plot_area.dTable.showKeys       = True
        except Exception:
            pass   # data table not critical if import fails

        wd.add_chart(chart, "K2")

        # ── STREAM ───────────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        now      = datetime.now()
        filename = (f"search-report_"
                    f"{now.strftime('%d').lower()}"
                    f"{now.strftime('%b').lower()}"
                    f"{now.strftime('%Y')}_"
                    f"{now.strftime('%H%M')}.xlsx")

        logger.info(f"Download ready — {filename} ({len(rows)} rows)")

        from flask import send_file
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Download error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@search_bp.route("/search/report")
def search_report():
    try:
        df, last_refresh = load_data()
        report = calculate_report(df)
        report["last_refresh"] = last_refresh
        logger.info(
            f"Report — incident={report['incident']['total']} "
            f"azure={report['azure']['total']} ptc={report['ptc']['total']} "
            f"azure_from_snow={len(report['azure_from_snow'])}"
        )
        return jsonify(report)
    except Exception as e:
        logger.error(f"Report error: {e}", exc_info=True)
        return jsonify({"error": str(e)})


@search_bp.route("/search/help-data")
def get_help_data():
    try:
        return jsonify({"sections": HELP_SECTIONS})
    except Exception as e:
        logger.error(f"Help data error: {e}", exc_info=True)
        return jsonify({"sections": [], "error": str(e)})
