import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Highlight styling configuration
FILL_ADDED    = PatternFill("solid", fgColor="C6EFCE")
FILL_MODIFIED = PatternFill("solid", fgColor="FFEB9C")
FILL_REMOVED  = PatternFill("solid", fgColor="FFC7CE")
FONT_REMOVED  = Font(color="9C0006", strike=True)

def normalize(val):
    if pd.isna(val): return ""
    return str(val).strip()

def load_all_sheets(path):
    # FORCE the engine to ensure 'xlsx' is supported
    engine = "xlrd" if path.lower().endswith(".xls") else "openpyxl"
    return pd.read_excel(path, sheet_name=None, header=None, engine=engine)

def diff_sheet(df1, df2):
    max_rows = max(len(df1), len(df2))
    max_cols = max(len(df1.columns), len(df2.columns))

    # Pad dataframes
    for c in range(len(df1.columns), max_cols): df1[c] = ""
    for c in range(len(df2.columns), max_cols): df2[c] = ""

    diff_mask1 = pd.DataFrame("", index=range(max_rows), columns=range(max_cols))
    diff_mask2 = pd.DataFrame("", index=range(max_rows), columns=range(max_cols))

    added_count = removed_count = modified_count = 0

    for r in range(max_rows):
        for c in range(max_cols):
            v1 = normalize(df1.iloc[r, c]) if r < len(df1) else ""
            v2 = normalize(df2.iloc[r, c]) if r < len(df2) else ""

            if r >= len(df1):
                diff_mask2.at[r, c] = "added"
                added_count += 1
            elif r >= len(df2):
                diff_mask1.at[r, c] = "removed"
                removed_count += 1
            elif v1 != v2:
                diff_mask2.at[r, c] = "modified"
                modified_count += 1

    summary = {"added": added_count, "removed": removed_count, "modified": modified_count}
    return diff_mask1, diff_mask2, summary
    
def build_side_by_side(df1, df2, diff_mask1, diff_mask2):
    rows = []
    max_rows = max(len(df1), len(df2))
    max_cols = max(len(df1.columns), len(df2.columns))

    for r in range(max_rows):
        left_cells, right_cells = [], []
        for c in range(max_cols):
            lval = normalize(df1.iloc[r, c]) if r < len(df1) and c < len(df1.columns) else ""
            lstatus = diff_mask1.at[r, c] if r < len(diff_mask1) and c < len(diff_mask1.columns) else ""
            rval = normalize(df2.iloc[r, c]) if r < len(df2) and c < len(df2.columns) else ""
            rstatus = diff_mask2.at[r, c] if r < len(diff_mask2) and c < len(diff_mask2.columns) else ""

            left_cells.append({"value": lval, "status": lstatus})
            right_cells.append({"value": rval, "status": rstatus})

        row_type = ""
        if any(cell["status"] == "removed" for cell in left_cells): row_type = "removed"
        elif any(cell["status"] == "added" for cell in right_cells): row_type = "added"
        elif any(cell["status"] == "modified" for cell in right_cells): row_type = "modified"

        rows.append({"row_num": r + 1, "type": row_type, "left": left_cells, "right": right_cells})
    return rows

def run_compare(path1, path2):
    sheets1 = load_all_sheets(path1)
    sheets2 = load_all_sheets(path2)
    all_sheets = list(dict.fromkeys(list(sheets1.keys()) + list(sheets2.keys())))
    results = {}
    t_added = t_removed = t_modified = 0

    for s_name in all_sheets:
        df1 = sheets1.get(s_name, pd.DataFrame()).reset_index(drop=True)
        df2 = sheets2.get(s_name, pd.DataFrame()).reset_index(drop=True)
        if df1.empty and df2.empty: continue

        dm1, dm2, summary = diff_sheet(df1, df2)
        sbs = build_side_by_side(df1, df2, dm1, dm2)

        t_added += summary["added"]; t_removed += summary["removed"]; t_modified += summary["modified"]

        log = []
        for r in range(len(df2)):
            for c in range(len(df2.columns)):
                status = dm2.at[r, c]
                if status in ("added", "modified"):
                    log.append({
                        "sheet": s_name, "cell": f"{chr(65+c)}{r+1}" if c < 26 else f"Col{c+1}R{r+1}",
                        "oldValue": normalize(df1.iloc[r, c]) if r < len(df1) and c < len(df1.columns) else "",
                        "newValue": normalize(df2.iloc[r, c]), "status": status.capitalize()
                    })

        results[s_name] = {
            "sbs": sbs, "change_log": log, "col_count": max(len(df1.columns), len(df2.columns)),
            "added": summary["added"], "removed": summary["removed"], "modified": summary["modified"]
        }

    return {
        "sheets": list(results.keys()), "sheet_data": results,
        "totals": {"added": t_added, "removed": t_removed, "modified": t_modified, "total": t_added + t_removed + t_modified}
    }