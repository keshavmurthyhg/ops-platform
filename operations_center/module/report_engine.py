# =============================================================================
#  REPORT ENGINE — Daily / Weekly / Summary / Range Reports
#  operations_center/module/report_engine.py
#
#  Outputs: HTML (in-browser), Excel (multi-sheet), PDF (print-ready)
# =============================================================================

import io
import json
import logging
from datetime import datetime, date, timedelta
from pathlib import Path

log = logging.getLogger("report_engine")

REPORT_BASE = Path("data") / "reports"

# ─────────────────────────────────────────────────────────────────────────────
#  Field helpers
# ─────────────────────────────────────────────────────────────────────────────

def _v(row, *keys):
    for k in keys:
        v = row.get(k, "")
        if v is not None and str(v).strip() not in ("", "nan", "None", "N/A"):
            return str(v).strip()
    return ""

def _safe(v, fallback="—"):
    s = str(v).strip() if v is not None else ""
    return s if s and s not in ("nan","None","N/A") else fallback

def _trunc(v, n=80):
    s = _safe(v)
    return s[:n]+"…" if len(s)>n else s

def _parse_date(val):
    if not val or str(val).strip() in ("","N/A","—","nan"): return None
    s = str(val).strip()
    for tz in (" CEST"," CET"," UTC"," GMT"): s = s.replace(tz,"")
    s = s.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S","%Y-%m-%d %H:%M","%Y-%m-%d",
                "%d-%b-%Y %H:%M","%d-%b-%Y","%d/%m/%Y","%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: pass
    return None

def _in_range(val, from_date, to_date):
    d = _parse_date(val)
    if d is None: return True  # undated = open item, include
    if from_date and d < from_date: return False
    if to_date   and d > to_date:   return False
    return True

def _filter(rows, date_keys, from_date, to_date):
    if not from_date and not to_date: return rows
    return [r for r in rows if _in_range(_v(r,*date_keys), from_date, to_date)]


# ─────────────────────────────────────────────────────────────────────────────
#  HTML Report Engine
# ─────────────────────────────────────────────────────────────────────────────

_STYLE = """<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Segoe UI,Arial,sans-serif;background:#f4f6f9;padding:20px;color:#374151;font-size:13px}
.wrap{max-width:1100px;margin:auto}
.report-header{background:linear-gradient(135deg,#1e3a5f,#1a2e4a);color:#fff;
  padding:20px 28px;border-radius:10px 10px 0 0}
.report-header h1{font-size:20px;font-weight:800;margin-bottom:4px}
.report-meta{font-size:11px;color:#93c5fd;display:flex;gap:18px;flex-wrap:wrap;margin-top:6px}
.report-body{background:#fff;border:1px solid #e5e7eb;border-top:none;padding:24px 28px}
.date-pill{display:inline-flex;align-items:center;gap:6px;background:#eff6ff;
  border:1px solid #bfdbfe;border-radius:6px;padding:5px 12px;
  font-size:12px;color:#1e40af;margin-bottom:18px}
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(110px,1fr));
  gap:10px;margin-bottom:22px}
.kpi{background:#f9fafb;border:1px solid #e5e7eb;border-radius:8px;
  padding:10px 14px;text-align:center}
.kpi.alert{border-color:#fca5a5;background:#fef2f2}
.kpi-val{font-size:22px;font-weight:800}
.kpi-lbl{font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;
  color:#6b7280;margin-top:2px}
.tracker-block{margin-bottom:28px;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden}
.tracker-hdr{display:flex;align-items:center;justify-content:space-between;
  padding:11px 16px;background:#f9fafb;border-bottom:1px solid #e5e7eb;flex-wrap:wrap;gap:8px}
.tracker-title{font-size:13px;font-weight:700;display:flex;align-items:center;gap:8px}
.tracker-chips{display:flex;gap:6px;flex-wrap:wrap}
.chip{display:inline-flex;flex-direction:column;align-items:center;padding:3px 10px;
  border-radius:6px;border:2px solid;min-width:50px;background:#fff;font-size:11px}
.chip-v{font-size:15px;font-weight:800}
.chip-l{font-size:8px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;opacity:.8}
.tracker-body{overflow-x:auto}
table{width:100%;border-collapse:collapse;font-size:12px}
th{background:#f9fafb;padding:7px 10px;text-align:left;font-weight:700;
   color:#374151;border-bottom:2px solid #e5e7eb;white-space:nowrap}
td{padding:6px 10px;border-bottom:1px solid #f3f4f6;vertical-align:top}
tr:hover td{background:#f9fafb}
.badge{display:inline-block;padding:2px 7px;border-radius:999px;font-size:10px;font-weight:700}
.red{background:#fef2f2;color:#dc2626}
.amber{background:#fffbeb;color:#d97706}
.green{background:#f0fdf4;color:#16a34a}
.blue{background:#eff6ff;color:#2563eb}
.gray{background:#f3f4f6;color:#6b7280}
.no-rec{color:#9ca3af;text-align:center;padding:16px;font-style:italic}
.report-footer{background:#f9fafb;border:1px solid #e5e7eb;border-top:none;
  border-radius:0 0 10px 10px;padding:10px 28px;font-size:11px;color:#9ca3af;
  display:flex;justify-content:space-between;flex-wrap:wrap;gap:6px}
.print-btn{position:fixed;bottom:20px;right:20px;background:#1e3a5f;color:#fff;
  border:none;border-radius:8px;padding:10px 18px;font-size:13px;font-weight:700;
  cursor:pointer;box-shadow:0 4px 16px rgba(0,0,0,.2);z-index:999}
@media print{.print-btn{display:none}body{padding:0;background:#fff}.wrap{max-width:100%}}
</style>"""

def _badge_html(val, show=True):
    if not show: return _safe(val)
    s = str(val).lower()
    if any(x in s for x in ("fail","error","hold","critical")): return f'<span class="badge red">{_safe(val)}</span>'
    if any(x in s for x in ("progress","pending","new","active","executing","open")): return f'<span class="badge amber">{_safe(val)}</span>'
    if any(x in s for x in ("success","complete","resolved","closed","ready")): return f'<span class="badge green">{_safe(val)}</span>'
    return f'<span class="badge gray">{_safe(val)}</span>'

def _chip(val, label, color):
    return (f'<div class="chip" style="color:{color};border-color:{color};">'
            f'<span class="chip-v">{val}</span>'
            f'<span class="chip-l">{label}</span></div>')

def _kpi(val, label, color, alert=False):
    return (f'<div class="kpi{"alert" if alert else ""}">'
            f'<div class="kpi-val" style="color:{color};">{val}</div>'
            f'<div class="kpi-lbl">{label}</div></div>')

def _table_html(rows, cols, show_badges=True, max_rows=500):
    head = "<thead><tr>" + "".join(f"<th>{c[0]}</th>" for c in cols) + "</tr></thead>"
    if not rows:
        return f"<table>{head}<tbody><tr><td colspan='{len(cols)}' class='no-rec'>No records in this date range.</td></tr></tbody></table>"
    body = "<tbody>"
    for r in rows[:max_rows]:
        body += "<tr>"
        for col in cols:
            keys = [k for k in col[1:] if not callable(k)]
            fn   = next((k for k in col[1:] if callable(k)), None)
            raw  = _v(r, *keys)
            if fn:
                body += f"<td>{fn(raw, r)}</td>"
            else:
                body += f"<td>{_safe(raw)}</td>"
        body += "</tr>"
    if len(rows) > max_rows:
        body += f"<tr><td colspan='{len(cols)}' style='color:#9ca3af;font-style:italic;padding:6px 10px;'>…and {len(rows)-max_rows} more rows</td></tr>"
    return f"<table>{head}{body}</tbody></table>"

def _tracker_block(title, color, icon, chips_html, table_html):
    return f"""<div class="tracker-block">
  <div class="tracker-hdr">
    <div class="tracker-title"><span>{icon}</span>{title}</div>
    <div class="tracker-chips">{chips_html}</div>
  </div>
  <div class="tracker-body">{table_html}</div>
</div>"""


def build_html_report(
    report_type, date_label, now,
    support_data, failure_data, incident_data,
    azure_data, ptc_data,
    transactions, wvs_queue, worker_stats,
    trackers, settings, from_date=None, to_date=None
):
    show_badges = settings.get("showBadgesInReport", True)
    title_map   = {
        "daily_report"  : "📋 Daily Operations Report",
        "weekly_report" : "📅 Weekly Operations Report",
        "summary_report": "📊 Summary Report — All Trackers",
        "range_report"  : "📆 Date Range Report",
    }
    title = title_map.get(report_type, "📋 Operations Report")
    logo  = settings.get("reportLogoText", "Ops Platform")
    frm   = settings.get("reportFromEmail","")
    to    = settings.get("reportToEmail","")
    T     = settings.get("alertThresholds", {})

    # ── Per-tracker sections ──────────────────────────────────────────────────
    sections = []

    if "failure" in trackers:
        rows     = _filter(failure_data, ["Failure Time","failure_time"], from_date, to_date)
        prod     = [r for r in rows if _v(r,"Environment","environment").upper()=="PROD"]
        tbl      = _table_html(
            sorted(rows, key=lambda r: _v(r,"Failure Time","failure_time"), reverse=True),
            [("Failure Time","Failure Time","failure_time"),
             ("Integration","Integration","Target","target"),
             ("Object Number","Object Number","Object","object"),
             ("Error Message","Error Message","Notes","notes",
              lambda v,r: f'<span title="{v}">{_trunc(v,90)}</span>'),
             ("Environment","Environment","environment",
              lambda v,r: _badge_html(v,show_badges) if v and v!="—" else "—"),
             ("WC Server","Windchill Server","wc_server")],
            max_rows=500)
        sections.append(_tracker_block(
            "Integration Failures","#ef4444","⚠️",
            _chip(len(rows),"Total","#ef4444") + _chip(len(prod),"PROD","#f59e0b"),
            tbl))

    if "support" in trackers:
        rows    = _filter(support_data, ["date_received","Date Received","Date"], from_date, to_date)
        pending = [r for r in rows if "Action Required" in _v(r,"category","Categories","Category")]
        tbl     = _table_html(
            pending or rows,
            [("Date","date_received","Date Received","Date"),
             ("From","name","Name"),
             ("Subject","subject","Subject",
              lambda v,r: f'<span title="{v}">{_trunc(v,100)}</span>'),
             ("Category","category","Categories","Category")],
            max_rows=500)
        sections.append(_tracker_block(
            "Support Emails","#3b82f6","📧",
            _chip(len(rows),"Total","#3b82f6") + _chip(len(pending),"Action Req","#ef4444"),
            tbl))

    if "incident" in trackers:
        on_hold = [r for r in incident_data if _v(r,"Status","status")=="On Hold"]
        in_prog = [r for r in incident_data if _v(r,"Status","status")=="In Progress"]
        tbl     = _table_html(incident_data,
            [("Number","Number","number"),
             ("Description","Description","short_description"),
             ("Assigned To","Assigned To","assigned_to"),
             ("Priority","Priority","priority"),
             ("Status","Status","status",lambda v,r:_badge_html(v,show_badges))],
            max_rows=500)
        sections.append(_tracker_block(
            "Incident Tracker","#f28c38","🎫",
            _chip(len(incident_data),"Total","#f28c38") +
            _chip(len(on_hold),"On Hold","#ef4444") +
            _chip(len(in_prog),"In Progress","#3b82f6"),
            tbl))

    if "azure" in trackers:
        new_    = [r for r in azure_data if _v(r,"Status","status")=="New"]
        active  = [r for r in azure_data if _v(r,"Status","status")=="Active"]
        tbl     = _table_html(azure_data,
            [("Number","Number","number"),
             ("Description","Description","title"),
             ("Assigned To","Assigned To","assigned_to"),
             ("Priority","Priority","priority"),
             ("Status","Status","status",lambda v,r:_badge_html(v,show_badges))],
            max_rows=500)
        sections.append(_tracker_block(
            "Azure Tracker","#0ea5e9","☁",
            _chip(len(azure_data),"Total","#0ea5e9") +
            _chip(len(new_),"New","#8b5cf6") +
            _chip(len(active),"Active","#22c55e"),
            tbl))

    if "ptc" in trackers:
        open_   = [r for r in ptc_data if _v(r,"Status","status") not in ("Closed","Resolved","Cancelled")]
        tbl     = _table_html(ptc_data,
            [("Number","Number","number"),
             ("Description","Description","subject"),
             ("Priority","Priority","severity","priority"),
             ("Status","Status","status",lambda v,r:_badge_html(v,show_badges))],
            max_rows=500)
        sections.append(_tracker_block(
            "PTC Cases","#8b5cf6","🛠",
            _chip(len(ptc_data),"Total","#8b5cf6") +
            _chip(len(open_),"Open","#f59e0b"),
            tbl))

    if "wm_tx" in trackers:
        tx_rows = _filter(transactions, ["time","Time"], from_date, to_date)
        failed  = [r for r in tx_rows if "FAIL" in _v(r,"status","Status").upper()]
        tbl     = _table_html(tx_rows,
            [("Time","time"),("Target","target"),("Action","action"),
             ("Status","status",lambda v,r:_badge_html(v,show_badges)),
             ("Object","object"),("State","state"),("Attempts","attempts"),
             ("Notes","notes",lambda v,r:f'<span title="{v}">{_trunc(v,60)}</span>')],
            max_rows=500)
        sections.append(_tracker_block(
            "WM — Transactions","#f28c38","⚡",
            _chip(len(tx_rows),"Total","#f28c38") +
            _chip(len(failed),"Failed","#ef4444"),
            tbl))

    if "wvs" in trackers:
        failed_wvs = [r for r in wvs_queue if "FAIL" in _v(r,"status","Status").upper()]
        tbl        = _table_html(wvs_queue,
            [("Queue","queue"),("Job","job"),
             ("Status","status",lambda v,r:_badge_html(v,show_badges)),
             ("Name","name"),("Version","version"),("User","user")],
            max_rows=500)
        sections.append(_tracker_block(
            "WM — WVS Queue","#0ea5e9","📋",
            _chip(len(wvs_queue),"Total","#0ea5e9") +
            _chip(len(failed_wvs),"Failed","#ef4444"),
            tbl))

    if "workers" in trackers:
        sorted_wk = sorted(worker_stats,
            key=lambda w: float(str(w.get("failed_pct","0")).replace("%","").strip() or "0"),
            reverse=True)
        def _pct(v,r):
            try:
                p=float(str(v).replace("%","").strip() or "0")
                c="#dc2626" if p>=T.get("wm_worker_fail_pct",10) else ("#d97706" if p>=5 else "#16a34a")
                return f'<strong style="color:{c};">{v}</strong>'
            except: return _safe(v)
        tbl = _table_html(sorted_wk,
            [("Worker","name"),("Total","total"),("Failed","failed"),
             ("Success","success"),("Fail %","failed_pct",_pct)],
            max_rows=500)
        sections.append(_tracker_block(
            "WM — Worker Stats","#374151","🖥",
            _chip(len(worker_stats),"Workers","#374151"),
            tbl))

    # ── Overall KPI row ───────────────────────────────────────────────────────
    all_fail  = _filter(failure_data, ["Failure Time","failure_time"], from_date, to_date)
    all_supp  = _filter(support_data, ["date_received","Date Received","Date"], from_date, to_date)
    all_tx    = _filter(transactions, ["time","Time"], from_date, to_date)
    kpi_html  = (
        _kpi(len(all_fail),   "Failures",     "#ef4444") +
        _kpi(len(all_supp),   "Support",      "#3b82f6") +
        _kpi(len(incident_data),"Incidents",  "#f28c38") +
        _kpi(len(azure_data), "Azure",        "#0ea5e9") +
        _kpi(len(ptc_data),   "PTC",          "#8b5cf6") +
        _kpi(len(all_tx),     "WM Tx",        "#f28c38") +
        _kpi(len(wvs_queue),  "WVS Jobs",     "#0ea5e9") +
        _kpi(len(worker_stats),"Workers",     "#374151")
    )

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">{_STYLE}
<title>{logo} — {title}</title></head>
<body><div class="wrap">
<div class="report-header">
  <h1>{title}</h1>
  <div class="report-meta">
    <span>🕐 {now}</span>
    <span>📤 {frm}</span>
    <span>📥 {to}</span>
  </div>
</div>
<div class="report-body">
  <div class="date-pill">📅 {date_label}</div>
  <div class="kpi-grid">{kpi_html}</div>
  {"".join(sections)}
</div>
<div class="report-footer">
  <span>{logo} · {title}</span>
  <span>{now}</span>
</div>
</div>
<button class="print-btn" onclick="window.print()">🖨 Print / Save PDF</button>
<script>
document.title = "{logo} — {date_label}";
</script>
</body></html>"""
    return html


# ─────────────────────────────────────────────────────────────────────────────
#  Excel multi-sheet report
# ─────────────────────────────────────────────────────────────────────────────

def build_excel_report(
    report_type, date_label, now,
    support_data, failure_data, incident_data,
    azure_data, ptc_data,
    transactions, wvs_queue, worker_stats,
    trackers, settings, from_date=None, to_date=None
):
    try:
        import openpyxl
        from openpyxl.styles import (PatternFill, Font, Alignment,
                                     Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl not installed — run: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Style helpers ─────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
    HDR_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ALT_FILL  = PatternFill("solid", fgColor="F9FAFB")
    THIN      = Side(style="thin", color="E5E7EB")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    TITLE_FONT= Font(bold=True, size=11, color="1E3A5F")

    def _add_sheet(name, headers, rows_data):
        ws = wb.create_sheet(title=name[:31])
        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=f"{name} — {date_label}")
        title_cell.font  = TITLE_FONT
        title_cell.fill  = PatternFill("solid", fgColor="EFF6FF")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 16

        # Header row
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font      = HDR_FONT
            c.fill      = HDR_FILL
            c.alignment = HDR_ALIGN
            c.border    = BORDER

        # Data rows
        for ri, row_vals in enumerate(rows_data, 3):
            fill = ALT_FILL if ri % 2 == 1 else None
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border    = BORDER
                if fill: c.fill = fill

        # Auto-width (capped at 60)
        for ci, h in enumerate(headers, 1):
            col_vals = [ws.cell(row=r, column=ci).value or "" for r in range(2, ws.max_row+1)]
            max_len  = max((len(str(v)) for v in col_vals if v), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 60)

        ws.freeze_panes = "A3"
        return ws

    T = settings.get("alertThresholds", {})

    if "failure" in trackers:
        rows = _filter(failure_data, ["Failure Time","failure_time"], from_date, to_date)
        rows = sorted(rows, key=lambda r: _v(r,"Failure Time","failure_time"), reverse=True)
        _add_sheet("Integration Failures",
            ["Failure Time","Integration","Object Number","Error Message","Environment","WC Server"],
            [[_v(r,"Failure Time","failure_time"),
              _v(r,"Integration","Target","target"),
              _v(r,"Object Number","Object","object"),
              _v(r,"Error Message","Notes","notes"),
              _v(r,"Environment","environment"),
              _v(r,"Windchill Server","wc_server")] for r in rows])

    if "support" in trackers:
        rows = _filter(support_data, ["date_received","Date Received","Date"], from_date, to_date)
        _add_sheet("Support Emails",
            ["Date Received","From","Subject","Category","Importance"],
            [[_v(r,"date_received","Date Received","Date"),
              _v(r,"name","Name"),
              _v(r,"subject","Subject"),
              _v(r,"category","Categories","Category"),
              _v(r,"importance","Importance")] for r in rows])

    if "incident" in trackers:
        _add_sheet("Incidents",
            ["Number","Description","Assigned To","Priority","Status","Vendor Ticket"],
            [[_v(r,"Number","number"),
              _v(r,"Description","short_description"),
              _v(r,"Assigned To","assigned_to"),
              _v(r,"Priority","priority"),
              _v(r,"Status","status"),
              _v(r,"Vendor Ticket","vendor_ticket")] for r in incident_data])

    if "azure" in trackers:
        _add_sheet("Azure Bugs",
            ["Number","Description","Assigned To","Priority","Status"],
            [[_v(r,"Number","number"),
              _v(r,"Description","title"),
              _v(r,"Assigned To","assigned_to"),
              _v(r,"Priority","priority"),
              _v(r,"Status","status")] for r in azure_data])

    if "ptc" in trackers:
        _add_sheet("PTC Cases",
            ["Number","Description","Priority","Status"],
            [[_v(r,"Number","number"),
              _v(r,"Description","subject"),
              _v(r,"Priority","severity","priority"),
              _v(r,"Status","status")] for r in ptc_data])

    if "wm_tx" in trackers:
        rows = _filter(transactions, ["time","Time"], from_date, to_date)
        _add_sheet("WM Transactions",
            ["Time","Target","Action","Status","Object","State","Attempts","Notes"],
            [[_v(r,"time"),_v(r,"target"),_v(r,"action"),_v(r,"status"),
              _v(r,"object"),_v(r,"state"),_v(r,"attempts"),_v(r,"notes")] for r in rows])

    if "wvs" in trackers:
        _add_sheet("WVS Queue",
            ["Queue","Job","Status","Name","Version","User"],
            [[_v(r,"queue"),_v(r,"job"),_v(r,"status"),
              _v(r,"name"),_v(r,"version"),_v(r,"user")] for r in wvs_queue])

    if "workers" in trackers:
        sorted_wk = sorted(worker_stats,
            key=lambda w: float(str(w.get("failed_pct","0")).replace("%","").strip() or "0"),
            reverse=True)
        _add_sheet("Worker Stats",
            ["Worker","Total","Failed","Success","Fail %"],
            [[_v(r,"name"),_v(r,"total"),_v(r,"failed"),
              _v(r,"success"),_v(r,"failed_pct")] for r in sorted_wk])

    # ── Summary sheet (first) ─────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15
    r = 1
    ws_sum.cell(r,1,"Ops Platform — "+now).font = Font(bold=True,size=12,color="1E3A5F"); r+=1
    ws_sum.cell(r,1,date_label).font = Font(italic=True,color="6B7280"); r+=2
    ws_sum.cell(r,1,"Tracker").font=Font(bold=True)
    ws_sum.cell(r,2,"Count").font=Font(bold=True); r+=1
    summaries = [
        ("Integration Failures", len(_filter(failure_data,["Failure Time","failure_time"],from_date,to_date))),
        ("Support Emails",        len(_filter(support_data,["date_received","Date Received"],from_date,to_date))),
        ("Incidents",             len(incident_data)),
        ("Azure Bugs",            len(azure_data)),
        ("PTC Cases",             len(ptc_data)),
        ("WM Transactions",       len(_filter(transactions,["time","Time"],from_date,to_date))),
        ("WVS Queue",             len(wvs_queue)),
        ("Workers",               len(worker_stats)),
    ]
    for name, cnt in summaries:
        ws_sum.cell(r,1,name); ws_sum.cell(r,2,cnt); r+=1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
