from flask import (
    Blueprint,
    render_template,
    jsonify,
    send_file,
    request
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from io import BytesIO
import pandas as pd  # <-- THIS WAS MISSING!
from datetime import datetime
from datetime import timedelta
import os


from dcn_analytics.module.services.dashboard_service import (
    load_dashboard_data
)

# Import your data file path variable from your validator module
from dcn_analytics.module.logic.validator import DATA_FILE

dcn_analytics_bp = Blueprint(
    "dcn_analytics",
    __name__
)

# Ensure you import your core logical engines for filters
from dcn_analytics.module.logic.validator import validate_master_dataset
from dcn_analytics.module.logic.analytics_engine import (
    prepare_dataframe,
    build_daily_summary,
    build_monthly_chart_data,
    build_monthly_pivot,
    build_kpi
)

# =========================================================
# PAGE
# =========================================================
@dcn_analytics_bp.route(
    "/dcn-analytics"
)
def dcn_analytics_page():

    return render_template(
        "dcn_analytics.html"
    )


# =========================================================
# DASHBOARD API
# =========================================================
@dcn_analytics_bp.route(
    "/api/dcn-analytics/dashboard"
)
def dashboard_api():

    try:

        result = load_dashboard_data()

        return jsonify(result)

    except Exception as error:

        return jsonify({

            "success": False,

            "message": str(error)

        })

# =========================================================
# UPLOAD EXCEL DATA FILE
# =========================================================
@dcn_analytics_bp.route(
    "/api/dcn-analytics/upload",
    methods=["POST"]
)
def upload_excel_api():
    try:
        if "file" not in request.files:
            return jsonify({
                "success": False,
                "message": "No file part in the request"
            }), 400

        file = request.files["file"]
        
        if file.filename == "":
            return jsonify({
                "success": False,
                "message": "No file selected"
            }), 400

        if file and (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
            # Ensure target directory exists
            os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
            
            # Save and overwrite master file
            file.save(DATA_FILE)
            
            return jsonify({
                "success": True,
                "message": "Master dataset uploaded and refreshed successfully!"
            })
            
        return jsonify({
            "success": False,
            "message": "Invalid file type. Please upload a valid Excel spreadsheet (.xlsx)"
        }), 400

    except Exception as error:
        return jsonify({
            "success": False,
            "message": f"Upload failed: {str(error)}"
        }), 500

# =========================================================
# DOWNLOAD DOWNLOAD REPORT: MONTHLY PIVOT
# =========================================================
@dcn_analytics_bp.route("/api/dcn-analytics/download/monthly")
def download_monthly_report():
    try:
        data = load_dashboard_data()
        pivot_records = data.get("monthly_pivot", [])
        
        df = pd.DataFrame(pivot_records)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Monthly Pivot Summary")
            
        output.seek(0)
        timestamp = datetime.now().strftime("%d%b%Y_%H%M")
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"Monthly_Pivot_Summary_{timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return str(e), 500


# =========================================================
# DOWNLOAD REPORT: DAILY SKIPPED SUMMARY
# =========================================================
@dcn_analytics_bp.route("/api/dcn-analytics/download/daily")
def download_daily_report():
    try:
        data = load_dashboard_data()
        daily_records = data.get("daily_summary", [])
        
        # Build clean export dataframe
        df = pd.DataFrame(daily_records)
        if not df.empty and "Total DCNs" in df.columns:
            df = df.drop(columns=["Total DCNs"])
            
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Daily Skipped Summary")
            
        output.seek(0)
        timestamp = datetime.now().strftime("%d%b%Y_%H%M")
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"Daily_Skipped_Summary_{timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return str(e), 500

# =========================================================
# DOWNLOAD REPORT: FULL EXCEL DASHBOARD
# =========================================================
@dcn_analytics_bp.route("/api/dcn-analytics/download/full-dashboard")
def download_full_excel_dashboard():
    try:
        data = load_dashboard_data()
        pivot_records = data.get("monthly_pivot", [])
        daily_records = data.get("daily_summary", [])

        wb = openpyxl.Workbook()
        
        # ─────────────────────────────────────────────────────────
        # STYLE DEFINITIONS (CALIBRI & CLEAN TABLES)
        # ─────────────────────────────────────────────────────────
        font_title  = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data   = Font(name="Calibri", size=11, color="000000")
        font_total  = Font(name="Calibri", size=11, bold=True, color="000000")
        
        # Professional Steel Blue Theme
        fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        fill_zebra  = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
        fill_total  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left   = Alignment(horizontal="left", vertical="center")
        
        thin_side   = Side(border_style="thin", color="D9D9D9")
        thick_top   = Side(border_style="thin", color="000000")
        double_bot  = Side(border_style="double", color="000000")
        
        border_data  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_total = Border(top=thick_top, bottom=double_bot, left=thin_side, right=thin_side)

        # ─────────────────────────────────────────────────────────
        # SHEET 1: DASHBOARD (PIVOT + LIVE CHART)
        # ─────────────────────────────────────────────────────────
        ws_dash = wb.active
        ws_dash.title = "Dashboard"
        ws_dash.views.sheetView[0].showGridLines = False  # Remove Grid lines

        # Row 2: Title Block
        ws_dash["B2"] = "DCN Analytics - Monthly Operations Dashboard"
        ws_dash["B2"].font = font_title
        ws_dash.row_dimensions[2].height = 28

        # Row 4: Pivot Headers
        headers = ["Month", "2023", "2024", "2025", "2026"]
        for col_idx, text in enumerate(headers, start=2): # Columns B to F
            cell = ws_dash.cell(row=4, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center if col_idx > 2 else align_left
            cell.border = border_data
        ws_dash.row_dimensions[4].height = 24

        # Rows 5-16: Monthly Data
        start_row = 5
        for idx, row in enumerate(pivot_records):
            curr_row = start_row + idx
            ws_dash.row_dimensions[curr_row].height = 20
            
            c_m = ws_dash.cell(row=curr_row, column=2, value=row.get("Month", "-"))
            c_3 = ws_dash.cell(row=curr_row, column=3, value=int(row.get("2023", 0)))
            c_4 = ws_dash.cell(row=curr_row, column=4, value=int(row.get("2024", 0)))
            c_5 = ws_dash.cell(row=curr_row, column=5, value=int(row.get("2025", 0)))
            c_6 = ws_dash.cell(row=curr_row, column=6, value=int(row.get("2026", 0)))
            
            c_m.alignment = align_left
            for cell in [c_m, c_3, c_4, c_5, c_6]:
                cell.font = font_data
                cell.border = border_data
                if cell != c_m:
                    cell.alignment = align_center
                if idx % 2 == 1: # Zebra striping
                    cell.fill = fill_zebra

        # Row 17: Summary Totals
        total_row = start_row + len(pivot_records)
        ws_dash.row_dimensions[total_row].height = 22
        t_label = ws_dash.cell(row=total_row, column=2, value="Total")
        t_label.font = font_total
        t_label.alignment = align_left
        t_label.fill = fill_total
        t_label.border = border_total
        
        for col_letter, col_idx in [("C", 3), ("D", 4), ("E", 5), ("F", 6)]:
            t_cell = ws_dash.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}5:{col_letter}{total_row-1})")
            t_cell.font = font_total
            t_cell.alignment = align_center
            t_cell.fill = fill_total
            t_cell.border = border_total

        # Live Editable Chart Configuration
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Monthly DCN Skipped Trend"
        chart.width = 16
        chart.height = 11
        chart_data = Reference(ws_dash, min_col=3, min_row=4, max_col=6, max_row=total_row-1)
        chart_cats = Reference(ws_dash, min_col=2, min_row=5, max_row=total_row-1)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(chart_cats)
        ws_dash.add_chart(chart, "H4")

        # Column Formatting Constraints
        ws_dash.column_dimensions["A"].width = 3
        ws_dash.column_dimensions["B"].width = 14
        for col in ["C", "D", "E", "F"]:
            ws_dash.column_dimensions[col].width = 11

        # Hide Unwanted Rows & Columns (Sheet 1)
        for col_idx in range(7, 30):  # Hide Columns G to AC (leave room for chart)
            if col_idx < 8 or col_idx > 24: # Keep chart space (H to X) visible
                ws_dash.column_dimensions[get_column_letter(col_idx)].hidden = True
        for row_idx in range(total_row + 2, 100):
            ws_dash.row_dimensions[row_idx].hidden = True

        # --------------------------------────────────────---------
        # SHEET 2: DAILY SKIPPED DETAILS (FORMATTED & FROZEN)
        # ---------------------------------------------------------
        ws_daily = wb.create_sheet(title="Daily Skipped Details")
        ws_daily.views.sheetView[0].showGridLines = False  # Remove Grid lines
        ws_daily.freeze_panes = "B2"  # Freeze row 1 (and column A padding spacer)

        # Row 1: Table Headers
        daily_headers = ["SL NO", "Date", "Sequence Skipped", "Skipped DCN Numbers"]
        for col_idx, text in enumerate(daily_headers, start=2): # Columns B to E
            cell = ws_daily.cell(row=1, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_left if col_idx == 5 else align_center
            cell.border = border_data
        ws_daily.row_dimensions[1].height = 24

        # Inject Rows
        for idx, row in enumerate(daily_records):
            r_idx = idx + 2
            ws_daily.row_dimensions[r_idx].height = 20
            
            c_sl   = ws_daily.cell(row=r_idx, column=2, value=idx + 1)
            c_date = ws_daily.cell(row=r_idx, column=3, value=row.get("Date", "-"))
            c_seq  = ws_daily.cell(row=r_idx, column=4, value=int(row.get("Sequence Skipped", 0)))
            c_nums = ws_daily.cell(row=r_idx, column=5, value=row.get("Skipped DCN Numbers", "-"))
            
            c_nums.alignment = align_left
            for cell in [c_sl, c_date, c_seq]:
                cell.alignment = align_center
                
            for cell in [c_sl, c_date, c_seq, c_nums]:
                cell.font = font_data
                cell.border = border_data
                if idx % 2 == 1:
                    cell.fill = fill_zebra

        # Dimension Calculations
        ws_daily.column_dimensions["A"].width = 3
        ws_daily.column_dimensions["B"].width = 10
        ws_daily.column_dimensions["C"].width = 15
        ws_daily.column_dimensions["D"].width = 18
        ws_daily.column_dimensions["E"].width = 50

        # Hide Unwanted Rows & Columns (Sheet 2)
        for col_idx in range(6, 30):  # Hide Column F onwards
            ws_daily.column_dimensions[get_column_letter(col_idx)].hidden = True
        for row_idx in range(len(daily_records) + 3, len(daily_records) + 100):
            ws_daily.row_dimensions[row_idx].hidden = True

        # ─────────────────────────────────────────────────────────
        # TRANSMIT COMPLETED STREAM
        # ─────────────────────────────────────────────────────────
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%d%b%Y_%H%M")
        return send_file(
            output, as_attachment=True,
            download_name=f"DCN_Operations_Dashboard_{timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as error:
        return jsonify({"success": False, "message": str(error)}), 500